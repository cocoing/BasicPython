#!/usr/local/bin/python3
# -*- coding: utf-8 -*-

# Now:P224页操作普查表，正在进行时
import openpyxl
import pprint
import os
os.chdir("/Users/hequan/Documents/BasicPython/census_pop")
wb = openpyxl.load_workbook("censuspopdata.xlsx")
ws = wb.active
countyData = {}

# fill in countyData with each county's population and tracts.
print("{:.^20}".format("Reading rows"))
for row in range(2, ws.max_row + 1):
    # Each row in the spreadsheet has data for one census tract.
    state = ws['B' + str(row)].value
    county = ws['C' + str(row)].value
    pop = ws['D' + str(row)].value

    # Open a new text file and write the contents of countyData to it.
    # Make sure the key for this state exists
    countyData.setdefault(state, {})
    # Make sure the key for this county in this state exists.
    countyData[state].setdefault(county, {'tracts': 0, 'pop': 0})
    # Each row represents one census tract, so increment by one.
    countyData[state][county]['tracts'] += 1
    # Increase the county pop by the pop in this census tract.
    countyData[state][county]['pop'] += int(pop)

# Open a new text file and write the contents of countyData to it.
print("writing results...")
with open("census2010.py", 'w') as resultFile:
    resultFile.write("allData" + pprint.pformat(countyData))
print("Done.")